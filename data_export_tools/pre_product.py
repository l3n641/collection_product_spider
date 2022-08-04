from openpyxl import load_workbook
import os


class PreProductExcel(object):
    columns = [
        "sku", "post_title", "post_content", "post_excerpt", "regular_price", "sale_price", "category",
        "tags", "featured_image", "pa_size", "stock", "post_status", "manage_stock", "stock_status",
        "backorders", "tax_status", "visibility", "featured", "comment_status", "ping_status", "menu_order",
        "post_author", "seopress_titles_title", "seopress_titles_desc", "post_date_gmt", "post_date",
    ]

    def __init__(self, path):
        self.path = path
        file_name = os.path.basename(self.path)
        self.file_name = file_name

    def get_product_data(self, sheet_name='Worksheet'):
        # 获取excel 产品数据
        try:
            workbook = load_workbook(filename=self.path, data_only=True)
            sheet = workbook[sheet_name]
            products = []
            for row in range(1, sheet.max_row + 1):
                data = {}
                for column_index, column_name in enumerate(self.columns):
                    data[column_name] = sheet.cell(row=row, column=column_index + 1).value

                products.append(data)

            return products
        except Exception as e:
            return False

    def update_product_category(self, category_list, sheet_name='Worksheet', export_path=None):

        export_path = export_path or self.path
        product_data = self.get_product_data()

        workbook = load_workbook(filename=self.path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            workbook.remove(worksheet)

        sheet = workbook.create_sheet(sheet_name, 0)

        row = 1
        for product in product_data:
            if product.get("category"):
                categories = product.get("category").split("|")
                validate_categories = []
                for category in categories:
                    if category in category_list:
                        validate_categories.append(category)

                new_categories = '|'.join(validate_categories)
                product["category"] = new_categories or product.get("category")

            for column_index, column_name in enumerate(self.columns):
                sheet.cell(row=row, column=column_index + 1).value = product.get(column_name)

            row = row + 1
        workbook.save(export_path)  # 保存
        workbook.close()  # 关闭
        return True
