from openpyxl import load_workbook
import os


class ProductExcel(object):

    def __init__(self, path):
        self.path = path
        file_name = os.path.basename(self.path)
        self.file_name = file_name

    def get_category(self, gender_column=1, cat_0_column=2, type_column=3, url_column=4, sheet_name='Sheet1'):
        # 1.打开 Excel
        try:
            workbook = load_workbook(filename=self.path, data_only=True)
            sheet = workbook[sheet_name]
            data = {}
            for row in range(1, sheet.max_row + 1):
                gender = self.column_name_filet(sheet.cell(row=row, column=gender_column).value, allow_nullable=True)
                cat_0 = self.column_name_filet(sheet.cell(row=row, column=cat_0_column).value)
                type_data = self.column_name_filet(sheet.cell(row=row, column=type_column).value,
                                                   allow_nullable=True) or ""
                url = sheet.cell(row=row, column=url_column).value.strip()
                category = f"{gender}->{cat_0}->{type_data}"
                data[url] = category
            return data
        except Exception as e:
            return False

    def write_product_list(self, product_detail_data, sheet_name='Sheet2'):
        try:
            workbook = load_workbook(filename=self.path)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                workbook.remove(worksheet)

            sheet = workbook.create_sheet(sheet_name)

            row = 1
            for item in product_detail_data:
                url = item.get("url")
                category = item.get("category")

                sheet.cell(row=row, column=1).value = url
                sheet.cell(row=row, column=2).value = category
                row = row + 1
            workbook.save(self.path)  # 保存
            workbook.close()  # 关闭
            return True
        except Exception as e:
            return False

    @staticmethod
    def get_category_by_url(category_dict, url):
        for key in category_dict:
            if url.startswith(key):
                return category_dict[key]

    @property
    def lange(self):
        file_name = os.path.basename(self.path)
        return file_name.split('.')[0].split("_")[-1].upper()

    def write_product_detail(self, product_detail_data, default_brand=None, sheet_name="Sheet4", export_path=None):

        columns = ["featured_image", "LANG", "CAT-0", "Category", "SIZE", "SKU", "Style-Name", "TITLE", "Brand",
                   "Brand-name", "model", "Type", "Gender", "Gender-name", "Color", "Color-Name", "desc", "desc2",
                   "price", "price2", "Description", "Keyword", "IMG-Add", "NPrice", "OPrice", "max", "min", "NSize",
                   "Date", "PageUrl"]
        export_path = export_path or self.path
        try:

            workbook = load_workbook(filename=self.path)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                workbook.remove(worksheet)

            sheet = workbook.create_sheet(sheet_name, 0)

            for i in range(len(columns)):
                sheet.cell(row=1, column=i + 1).value = columns[i]

            row = 2
            for product in product_detail_data:
                if not product.featured_image or not product.title:
                    continue
                size = "|".join(set(product.size.split("|"))) if product.size else None
                size = ''
                if product.size:
                    size_set = []
                    for size_data in product.size.split("|"):
                        if size_data not in size_set:
                            size_set.append(size_data)

                    size = '|'.join(size_set)
                data = [
                    product.featured_image,
                    self.lange,
                    product.cat_0,
                    product.category_name,
                    size,
                    product.sku,
                    product.title,
                    product.title,
                    product.get_brand(default_brand),
                    product.get_brand(default_brand),
                    "",
                    product.category_type,
                    product.gender,
                    product.gender,
                    product.color,
                    product.color,
                    product.basc,
                    product.basc,
                    product.price,
                    product.price,
                    "", "", "", "", "", "", "", "",
                    product.dade,
                    product.PageUrl, ]
                for i in range(len(columns)):
                    sheet.cell(row=row, column=i + 1).value = data[i]
                row = row + 1
            workbook.save(export_path)  # 保存
            workbook.close()  # 关闭
            return True
        except Exception as e:
            return False

    @property
    def project_name(self):
        file_name = os.path.basename(self.path).split(".")[0]
        return file_name

    @staticmethod
    def column_name_filet(name, allow_nullable=False):
        if not allow_nullable and (not name or not name.strip()):
            raise ValueError("字段不能为空")
        return name.strip().title().replace("|", "&") if name else ""
